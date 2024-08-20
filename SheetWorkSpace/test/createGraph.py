import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from math import pi
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.font_manager as fm
import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox

# 맑은 고딕 폰트 파일 경로
font_path = 'C:\\Windows\\Fonts\\malgun.ttf'
font_prop = fm.FontProperties(fname=font_path)
plt.rcParams['font.family'] = font_prop.get_name()

# 시트 제목에서 허용되지 않는 문자를 제거하는 함수
def sanitize_sheet_title(title):
    if isinstance(title, bytes):
        title = title.decode('utf-8')
    invalid_chars = re.compile(r'[\\/*?[\]:]')
    sanitized_title = invalid_chars.sub('', title)
    if len(sanitized_title) > 31:
        sanitized_title = sanitized_title[:31]  # 엑셀 시트 이름은 최대 31자
    return sanitized_title

# GUI 코드 시작
def process_files(my_view_path, others_view_path, output_path, image_dir):
    try:
        # 'my_view.xlsx'와 평균값이 저장된 'others_view' 파일을 읽어들입니다.
        df_my_view = pd.read_excel(my_view_path)
        df_others_avg = pd.read_excel(others_view_path)

        # 데이터프레임의 컬럼 이름을 확인합니다.
        print("df_my_view columns:", df_my_view.columns)
        print("df_others_avg columns:", df_others_avg.columns)
    except Exception as e:
        messagebox.showerror("파일 오류", f"파일을 읽는 데 오류가 발생했습니다: {e}")
        return

    # 두 데이터프레임을 학번을 기준으로 병합합니다.
    df_merged = pd.merge(df_my_view, df_others_avg, on=['이름', '학번'], suffixes=('_내가', '_남이'))

    # 데이터프레임의 컬럼 이름을 확인합니다.
    print("df_merged columns:", df_merged.columns)

    # 카테고리 목록
    categories = df_my_view.columns[2:]  # '이름'과 '학번'을 제외한 나머지
    num_vars = len(categories)

    # 이미지 저장 경로를 설정합니다.
    os.makedirs(image_dir, exist_ok=True)

    # 엑셀 워크북을 생성합니다.
    wb = Workbook()
    default_sheet = wb.active

    # 기본 시트가 존재하면 삭제합니다.
    if default_sheet is not None:
        wb.remove(default_sheet)

    # 적어도 하나의 시트가 반드시 있어야 합니다.
    has_visible_sheet = False

    # 각 행에 대해 방사형 그래프를 생성하고 시트에 추가합니다.
    for index, row in df_merged.iterrows():
        name = row['이름']
        student_id = row['학번']

        try:
            # 방사형 그래프 데이터를 추출합니다.
            my_view = row[[f'{cat}_내가' for cat in categories]].values.flatten().tolist()
            others_view = row[[f'{cat}_남이' for cat in categories]].values.flatten().tolist()
        except KeyError as e:
            messagebox.showerror("컬럼 오류", f"컬럼을 찾을 수 없습니다: {e}")
            continue

        # 방사형 그래프를 생성합니다.
        angles = [n / float(num_vars) * 2 * pi for n in range(num_vars)]
        my_view += my_view[:1]  # 첫 번째 값을 추가하여 원을 완성
        others_view += others_view[:1]  # 첫 번째 값을 추가하여 원을 완성
        angles += angles[:1]  # 첫 번째 값을 추가하여 원을 완성

        fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))

        ax.set_theta_offset(pi / 2)
        ax.set_theta_direction(-1)

        plt.xticks(angles[:-1], categories, fontproperties=font_prop)

        def add_values_to_plot(angles, values, color):
            for i in range(len(values) - 1):
                angle = angles[i]
                value = values[i]
                x = angle
                y = value
                ax.text(x, y + 0.1, str(value), color=color, ha='center', va='center', fontsize=8, fontproperties=font_prop)

        ax.plot(angles, my_view, linewidth=2, linestyle='solid', color='blue', label='내가 보는 나의 모습')
        ax.fill(angles, my_view, color='blue', alpha=0.25)
        add_values_to_plot(angles, my_view, 'blue')

        ax.plot(angles, others_view, linewidth=2, linestyle='solid', color='red', label='남이 보는 나의 모습')
        ax.fill(angles, others_view, color='red', alpha=0.1)
        add_values_to_plot(angles, others_view, 'red')

        plt.title(name, size=20, color='blue', y=1.1, fontproperties=font_prop)
        plt.legend(loc='upper right', bbox_to_anchor=(0.1, 0.1), prop=font_prop)

        # 그래프를 이미지로 저장합니다.
        image_path = os.path.join(image_dir, f'{student_id}.png')
        plt.savefig(image_path, bbox_inches='tight')
        plt.close(fig)

        # 새로운 시트를 생성하고 이미지 추가합니다.
        sanitized_name = sanitize_sheet_title(str(student_id))
        ws = wb.create_sheet(title=sanitized_name)

        img = Image(image_path)
        ws.add_image(img, 'A1')

        # 데이터를 전치하여 엑셀에 표 형태로 추가합니다.
        data_with_id = pd.DataFrame({
            '카테고리': categories,
            '내가 보는 나의 모습': my_view[:-1],
            '남이 보는 나의 모습': others_view[:-1]
        }).T

        # 표를 시트에 추가합니다.
        start_row = 30  # 그래프 아래에 표를 시작할 위치
        for r_idx, row in enumerate(data_with_id.values, start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        has_visible_sheet = True

    if not has_visible_sheet:
        wb.create_sheet(title='Empty')

    try:
        # 엑셀 파일로 저장합니다.
        wb.save(output_path)
        messagebox.showinfo("완료", f"방사형 그래프와 표가 저장된 파일: {output_path}")
    except Exception as e:
        messagebox.showerror("저장 오류", f"파일을 저장하는 데 오류가 발생했습니다: {e}")

# 파일 선택을 위한 함수
def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    return file_path

# GUI 구성
root = tk.Tk()
root.withdraw()  # 기본 GUI 창 숨기기

# 'my_view.xlsx' 파일 선택
messagebox.showinfo("파일 선택", "내가 보는 나 파일을 선택하세요.")
my_view_path = select_file()

# 'others_view.xlsx' 파일 선택
messagebox.showinfo("파일 선택", "평균을 낸 파일을 선택하세요.")
others_view_path = select_file()

# 출력 엑셀 파일 경로 선택
messagebox.showinfo("파일 저장", "저장할 엑셀 파일의 경로를 선택하세요.")
output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

# 이미지 저장 디렉토리 선택
messagebox.showinfo("디렉토리 선택", "이미지를 저장할 디렉토리를 선택하세요.")
image_dir = filedialog.askdirectory()

# 선택된 파일 경로 및 디렉토리를 사용하여 파일 처리
process_files(my_view_path, others_view_path, output_path, image_dir)
